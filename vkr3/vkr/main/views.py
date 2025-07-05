from django.shortcuts import get_object_or_404
from django.shortcuts import render, redirect
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import user_passes_test
from .forms import RegisterForm
from django.contrib import messages
from .forms import ProfileUpdateForm
from django.urls import reverse_lazy
from django.contrib.auth.views import LoginView
from .forms import CourseCreateForm, CourseEnrollForm, CurriculumUploadForm
from .models import Course, CourseContent, CourseFile, Curriculum, CurriculumItem, User, Group, GroupDisciplineAssignment, StudentWork
from .forms import CourseContentForm
from django.http import FileResponse, Http404, HttpResponseForbidden
from django.core.exceptions import PermissionDenied
from .curriculum_parser import parse_curriculum
from collections import defaultdict
from .forms import AssignTeacherForm
from django.utils.text import slugify
from django.urls import reverse
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.utils.text import slugify



@login_required

def profile_edit(request):
    if request.method == "POST":
        form = ProfileUpdateForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Данные успешно обновлены.")
            return redirect("main:profile")
    else:
        form = ProfileUpdateForm(instance=request.user)

    return render(request, "main/profile_edit.html", {"form": form})

def profile_view(request):
    user = request.user
    student_courses = user.enrolled_courses.all() if user.role == 'student' else []
    teacher_courses = user.created_courses.all() if user.role == 'teacher' or user.role == 'admin' else []

    context = {
        'user': user,
        'student_courses': student_courses,
        'teacher_courses': teacher_courses,
    }
    return render(request, 'main/profile.html', context)
def home(request):
    return render(request, 'main/home.html')

def register(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)  # Выполнить вход сразу после регистрации
            return redirect('main:home')  # Перенаправление на главную страницу
    else:
        form = RegisterForm()
    return render(request, 'main/register.html', {'form': form})

class CustomLoginView(LoginView):
    template_name = "main/login.html"
    redirect_authenticated_user = True

    def get_success_url(self):
        return reverse_lazy("main:profile")  # Перенаправление после входа

#Курсы
@login_required
def course_list(request):
    query = request.GET.get("q", "")
    courses = Course.objects.filter(title__icontains=query) if query else Course.objects.all()

    created_courses = []
    enrolled_courses = []

    if request.user.is_authenticated:
        if request.user.role == "teacher" or request.user.is_superuser:
            created_courses = Course.objects.filter(teacher=request.user)
        elif request.user.role == "student":
            enrolled_courses = request.user.enrolled_courses.all()

    return render(request, "main/courses.html", {
        "courses": courses,
        "created_courses": created_courses,
        "enrolled_courses": enrolled_courses,
    })

@login_required
def create_course(request):
    if request.user.role != "teacher" and not request.user.is_superuser:
        return redirect("main:course_list")

    if request.method == "POST":
        form = CourseCreateForm(request.POST)
        if form.is_valid():
            course = form.save(commit=False)
            course.teacher = request.user
            course.save()
            return redirect("main:course_list")
    else:
        form = CourseCreateForm()

    return render(request, "main/create_course.html", {"form": form})

@login_required
def enroll_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    if request.user.role == "student":
        course.students.add(request.user)
    return redirect("main:course_list")

# Содержание курса

@login_required
def course_detail(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    contents = course.contents.all()
    return render(request, "main/course_detail.html", {"course": course, "contents": contents})


@login_required
def add_course_content(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    # Проверка прав пользователя
    if request.user != course.teacher and not request.user.is_superuser:
        messages.error(request, "У вас нет прав на добавление материала.")
        return redirect("main:course_detail", course_id=course_id)

    if request.method == "POST":
        form = CourseContentForm(request.POST, request.FILES)
        if form.is_valid():
            # Сохраняем новый контент курса
            content = form.save(commit=False)
            content.course = course
            content.save()

            # Если были загружены файлы, сохраняем их в модель CourseFile
            if request.FILES.getlist('file'):
                for uploaded_file in request.FILES.getlist('file'):
                    CourseFile.objects.create(content=content, file=uploaded_file)

            messages.success(request, "Материал успешно добавлен!")
            return redirect("main:course_detail", course_id=course.id)
    else:
        form = CourseContentForm()

    return render(request, "main/add_course_content.html", {
        "form": form,
        "course": course
    })
def download_file(request, course_id, file_id):
    material = get_object_or_404(CourseContent, id=file_id, course_id=course_id)

    if not material.file:
        raise Http404("Файл не найден")

    return FileResponse(material.file.open('rb'), as_attachment=True)

# Изменение и удаление содержимого курса

@login_required
def edit_course_material(request, course_id, material_id):
    course = get_object_or_404(Course, id=course_id)
    material = get_object_or_404(CourseContent, id=material_id, course=course)

    if request.user != course.teacher and request.user.role != "admin":
        messages.error(request, "Вы не имеете прав на редактирование этого материала.")
        return redirect("main:course_detail", course_id=course.id)

    if request.method == "POST":
        form = CourseContentForm(request.POST, request.FILES, instance=material)
        if form.is_valid():
            # Сохраняем изменения текста материала
            content = form.save()

            # Если загружены новые файлы, добавляем их в модель CourseFile
            if request.FILES.getlist('file'):
                for uploaded_file in request.FILES.getlist('file'):
                    CourseFile.objects.create(content=content, file=uploaded_file)

            messages.success(request, "Материал успешно обновлён!")
            return redirect("main:course_detail", course_id=course.id)
    else:
        form = CourseContentForm(instance=material)

    # Получаем прикрепленные файлы для отображения
    course_files = CourseFile.objects.filter(content=material)

    return render(request, "main/edit_material.html", {
        "form": form,
        "course": course,
        "material": material,
        "course_files": course_files
    })

@login_required
def delete_course_material(request, course_id, material_id):
    course = get_object_or_404(Course, id=course_id)
    material = get_object_or_404(CourseContent, id=material_id, course=course)

    if request.user != course.teacher and request.user.role != "admin":
        messages.error(request, "Вы не имеете прав на удаление этого материала.")
        return redirect("main:course_detail", course_id=course.id)

    material.delete()
    messages.success(request, "Материал удалён.")
    return redirect("main:course_detail", course_id=course.id)

# Удаление и изменение курса

@login_required
def edit_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    # Проверяем, имеет ли пользователь права на редактирование
    if not (request.user.is_superuser or (request.user.role == "teacher" and request.user == course.teacher)):
        raise PermissionDenied

    if request.method == "POST":
        form = CourseCreateForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            return redirect("main:course_detail", course_id=course.id)
    else:
        form = CourseCreateForm(instance=course)

    return render(request, "main/edit_course.html", {"form": form, "course": course})

@login_required
def delete_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    # Проверяем права на удаление
    if not (request.user.is_superuser or (request.user.role == "teacher" and request.user == course.teacher)):
        raise PermissionDenied

    course.delete()
    return redirect("main:course_list")  # Перенаправляем на список курсов

#Учебный план
@login_required
def curriculum_list_view(request):
    if not request.user.is_superuser:
        return HttpResponseForbidden()

    plans = Curriculum.objects.all().order_by('-created_at')
    return render(request, "main/curriculum_list.html", {"plans": plans})
def deep_convert(d):
    if isinstance(d, defaultdict):
        d = {k: deep_convert(v) for k, v in d.items()}
    return d
@login_required
def curriculum_detail_view(request, curriculum_id):
    curriculum = get_object_or_404(Curriculum, id=curriculum_id)
    return render(request, 'curriculums/detail.html', {'curriculum': curriculum})
@login_required
def curriculum_view(request, curriculum_id=None):
    user = request.user

    if curriculum_id:
        curriculum = get_object_or_404(Curriculum, id=curriculum_id)
        if not user.is_superuser:
            if not user.group or user.group.curriculum.id != curriculum.id:
                return HttpResponseForbidden("Нет доступа.")
    else:
        curriculum = user.group.curriculum

    # Получаем предметы
    items = CurriculumItem.objects.filter(curriculum=curriculum).select_related('subject')

    # Получаем преподавателей, назначенных на предметы для конкретной группы
    teacher_map = {}
    if user.group:
        assignments = GroupDisciplineAssignment.objects.filter(
            group=user.group,
            curriculum_item__in=items
        ).select_related('teacher', 'curriculum_item')

        for a in assignments:
            teacher_map[a.curriculum_item.id] = a.teacher

    # Группируем по курсам и семестрам
    course_semesters = defaultdict(lambda: defaultdict(list))
    max_course = user.group.course_year if hasattr(user.group, 'course_year') else None

    for item in items:
        if user.is_superuser or (max_course and item.course <= max_course):
            teacher = teacher_map.get(item.id)
            course_semesters[item.course][item.semester].append({
                'item': item,
                'teacher': teacher,
            })

    sorted_course_semesters = dict(sorted(course_semesters.items()))  # Сортируем по курсу (1, 2, 3, ...)
    for course in sorted_course_semesters:
        sorted_course_semesters[course] = dict(sorted(sorted_course_semesters[course].items()))  # Сортируем семестры

    return render(request, "main/curriculum.html", {
        "curriculum": curriculum,
        "is_admin": user.is_superuser,
        "course_semesters": deep_convert(sorted_course_semesters),
    })
@login_required
@user_passes_test(lambda u: u.is_superuser)
def delete_curriculum_view(request, curriculum_id):
    curriculum = get_object_or_404(Curriculum, id=curriculum_id)
    curriculum.delete()
    messages.success(request, "Учебный план успешно удален.")
    return redirect('main:curriculum_list')

@login_required
def upload_curriculum(request):
    if not request.user.is_superuser:
        return redirect("curriculum")

    if request.method == "POST":
        form = CurriculumUploadForm(request.POST, request.FILES)
        if form.is_valid():
            curriculum = form.save(commit=False)
            curriculum.uploaded_by = request.user
            curriculum.save()

            # Парсинг и сохранение
            parse_curriculum(curriculum.file.path, curriculum)

            return redirect("main:curriculum_list")
    else:
        form = CurriculumUploadForm()

    return render(request, "main/upload_curriculum.html", {"form": form})

@user_passes_test(lambda u: u.is_superuser)
def teacher_list_view(request):
    teachers = User.objects.filter(role__in=['teacher', 'admin'])
    return render(request, 'main/teacher_list.html', {'teachers': teachers})

@user_passes_test(lambda u: u.is_superuser)
def assign_subjects_view(request, teacher_id):
    teacher = get_object_or_404(User, id=teacher_id)
    groups = Group.objects.select_related('curriculum').all()
    assignments = []

    for group in groups:
        items = CurriculumItem.objects.filter(curriculum=group.curriculum).select_related('subject')
        for item in items:
            assignment, _ = GroupDisciplineAssignment.objects.get_or_create(
                group=group, curriculum_item=item
            )
            assignments.append(assignment)
    assignments.sort(key=lambda a: (a.group.name, a.curriculum_item.course, a.curriculum_item.semester))
    if request.method == "POST":
        selected_ids = request.POST.getlist('assignments')
        for a in assignments:
            if str(a.id) in selected_ids:
                a.teacher = teacher
            elif a.teacher == teacher:
                a.teacher = None
            a.save()
        return redirect('main:teacher_list')

    return render(request, 'main/assign_subjects.html', {
        'teacher': teacher,
        'assignments': assignments,
    })

@login_required
def teacher_subjects_view(request):
    user = request.user

    if user.role not in ['teacher', 'admin']:
        return HttpResponseForbidden("Доступ запрещен.")

    # === POST-запрос: обновление количества лаб/практик ===
    if request.method == 'POST':
        assignment_id = request.POST.get('assignment_id')
        update_type = request.POST.get('update_type')

        try:
            a = GroupDisciplineAssignment.objects.get(id=assignment_id, teacher=user)
            old_lab_count = a.lab_count or 0
            old_practice_count = a.practice_count or 0

            # Лабораторные
            if update_type in ['labs', 'both'] and a.curriculum_item.has_laboratory:
                val = request.POST.get(f'labs_{a.id}')
                if val and val.isdigit():
                    a.lab_count = int(val)
                else:
                    messages.error(request, "Неверное значение лабораторных.")

            # Практики
            if update_type in ['practices', 'both'] and a.curriculum_item.has_practice:
                val = request.POST.get(f'practices_{a.id}')
                if val and val.isdigit():
                    a.practice_count = int(val)
                else:
                    messages.error(request, "Неверное значение практик.")

            a.lab_count = a.lab_count or 0
            a.practice_count = a.practice_count or 0
            a.save()

            # Удаление лишних StudentWork, если уменьшилось количество
            if a.lab_count < old_lab_count:
                StudentWork.objects.filter(assignment=a, work_type='lab', lab_index__gte=a.lab_count).delete()
            if a.practice_count < old_practice_count:
                StudentWork.objects.filter(assignment=a, work_type='practice', lab_index__gte=a.practice_count).delete()

            messages.success(request, f"Обновлены данные для: {a.curriculum_item.subject.name}")

        except GroupDisciplineAssignment.DoesNotExist:
            messages.error(request, "Назначение не найдено.")

        return redirect(reverse('main:teacher_subjects'))

    # === GET-запрос ===
    assignments = GroupDisciplineAssignment.objects.filter(
        teacher=user
    ).select_related('curriculum_item__subject', 'group').prefetch_related('group__students')

    student_works = StudentWork.objects.filter(
        assignment__in=assignments
    ).select_related('student')

    # (assignment_id, student_id) → list of works
    student_work_map = defaultdict(list)
    for w in student_works:
        # Приводим assignment_id и student_id к int для консистентности ключей
        student_work_map[(int(w.assignment_id), int(w.student_id))].append(w)

    # Подготовка данных
    for a in assignments:
        item = a.curriculum_item
        a.subject_name = item.subject.name
        a.has_labs = item.has_laboratory
        a.has_practice = item.has_practice
        a.lab_count_display = a.lab_count or 0
        a.practice_count_display = a.practice_count or 0
        a.lab_range = range(a.lab_count_display)
        a.practice_range = range(a.practice_count_display)
        a.work_types = item.work_types.split(', ') if item.work_types else []
        a.control_types = item.control_types.split(', ') if item.control_types else []

        students = list(a.group.students.all())
        for student in students:
            key = (a.id, student.id)
            works = student_work_map.get(key, [])

            # Используем lower() для сравнения и ключи lab_index как строки для консистентности
            student.lab_works = {str(w.lab_index): w for w in works if w.work_type.lower() == "lab"}
            student.practice_works = {str(w.lab_index): w for w in works if w.work_type.lower() == "practice"}
            student.work_grades = {w.work_type: w for w in works if w.work_type.lower() not in ["lab", "practice"]}


        a.students = students

        # === Расчёт процента успеваемости ===
        total = 0
        done = 0

        for student in students:
            # Лабораторные
            for i in a.lab_range:
                total += 1
                if str(i) in student.lab_works and student.lab_works[str(i)].is_completed:
                    done += 1

            # Практики
            for i in a.practice_range:
                total += 1
                practice = student.practice_works.get(str(i))
                if practice and practice.grade:
                    done += 1

            # Прочие виды работ
            for wt in a.work_types:
                total += 1
                w = student.work_grades.get(wt)
                if w and w.grade:
                    done += 1

            # Контрольные формы
            for ct in a.control_types:
                total += 1
                c = student.work_grades.get(ct)
                if c and (c.grade or c.is_passed):
                    done += 1

        a.progress_percent = round((done / total) * 100) if total else 0

    return render(request, 'main/teacher_subjects.html', {
        'assignments': assignments,
    })

@login_required
def grade_students_view(request, assignment_id):
    assignment = get_object_or_404(GroupDisciplineAssignment, id=assignment_id, teacher=request.user)
    students = assignment.group.students.all()

    if request.method == "POST":
        for student in students:
            # Лабораторные
            if assignment.curriculum_item.has_laboratory:
                for i in range(assignment.lab_count or 0):
                    is_done = f"lab_{student.id}_{i}" in request.POST
                    StudentWork.objects.update_or_create(
                        student=student,
                        assignment=assignment,
                        work_type="lab",
                        lab_index=i,
                        defaults={"is_completed": is_done}
                    )

            # Практики
            if assignment.curriculum_item.has_practice:
                for i in range(assignment.practice_count or 0):
                    key = f"practice_{student.id}_{i}"
                    value = request.POST.get(key)
                    if value and value.isdigit():
                        StudentWork.objects.update_or_create(
                            student=student,
                            assignment=assignment,
                            work_type="practice",
                            lab_index=i,
                            defaults={"grade": int(value)}
                        )

            # Остальные виды работ
            types = (assignment.curriculum_item.work_types or "").split(', ')
            for wt in types:
                key = f"work_{slugify(wt)}_{student.id}"
                value = request.POST.get(key)
                if value and value.isdigit():
                    StudentWork.objects.update_or_create(
                        student=student,
                        assignment=assignment,
                        work_type=wt,
                        defaults={"grade": int(value)}
                    )

            # Контрольные формы
            controls = (assignment.curriculum_item.control_types or "").split(', ')
            for ct in controls:
                key = f"control_{slugify(ct)}_{student.id}"
                value = request.POST.get(key)
                if value != "":
                    if ct == "зачет":
                        StudentWork.objects.update_or_create(
                            student=student,
                            assignment=assignment,
                            work_type=ct,
                            defaults={"is_passed": value == "1"}
                        )
                    elif value.isdigit():
                        StudentWork.objects.update_or_create(
                            student=student,
                            assignment=assignment,
                            work_type=ct,
                            defaults={"grade": int(value)}
                        )

        return redirect("main:teacher_subjects")

@login_required
def export_student_works_excel(request, assignment_id):
    try:
        assignment = GroupDisciplineAssignment.objects.select_related(
            'curriculum_item__subject', 'group'
        ).prefetch_related('group__students').get(id=assignment_id, teacher=request.user)

        students = assignment.group.students.all()
        works = StudentWork.objects.filter(assignment=assignment).select_related('student')

        # Организуем работы по студенту
        work_map = defaultdict(lambda: defaultdict(dict))
        for w in works:
            key = w.lab_index if w.work_type in ['lab', 'practice'] else None
            work_map[w.student_id][w.work_type][key] = w

        # Структура
        lab_range = range(assignment.lab_count or 0)
        practice_range = range(assignment.practice_count or 0)
        work_types = assignment.curriculum_item.work_types.split(', ') if assignment.curriculum_item.work_types else []
        control_types = assignment.curriculum_item.control_types.split(', ') if assignment.curriculum_item.control_types else []

        # Создаем книгу
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ведомость"

        # === 1. Заголовок дисциплины ===
        discipline_name = f"{assignment.curriculum_item.subject.name} — {assignment.group.name}, " \
                          f"{assignment.curriculum_item.course} курс, {assignment.curriculum_item.semester} семестр"

        # Добавляем первую строку с заголовком
        ws.append([discipline_name])

        # Заголовки
        headers = ["ФИО студента"]
        headers += [f"Лаб. {i+1}" for i in lab_range]
        headers += [f"Пр. {i+1}" for i in practice_range]
        headers += work_types
        headers += control_types

        # Объединяем ячейки первой строки по ширине всей таблицы
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws["A1"].font = Font(bold=True, size=12)
        ws["A1"].alignment = Alignment(horizontal="center")

        # === 2. Заголовки таблицы ===
        ws.append(headers)


        # Данные по студентам
        for student in students:
            row = [student.get_full_name()]
            s_works = work_map.get(student.id, {})

            # Лабы
            for i in lab_range:
                w = s_works.get("lab", {}).get(i)
                row.append("сдано" if w and w.is_completed else "")

            # Практики
            for i in practice_range:
                w = s_works.get("practice", {}).get(i)
                row.append(w.grade if w and w.grade else "")

            # Прочие работы
            for wt in work_types:
                w = s_works.get(wt, {}).get(None)
                row.append(w.grade if w and w.grade else "")

            # Контрольные
            for ct in control_types:
                w = s_works.get(ct, {}).get(None)
                if ct.lower() in ['зачет', 'зачет с оценкой']:
                    val = "зачтено" if w and w.is_passed else "не зачтено" if w and w.is_passed is False else ""
                else:
                    val = w.grade if w and w.grade else ""
                row.append(val)

            ws.append(row)

        # Ширина колонок
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        # Ответ
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"student_works_{assignment.curriculum_item.subject.name}_{assignment.group.name}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    except GroupDisciplineAssignment.DoesNotExist:
        return HttpResponse("Дисциплина не найдена", status=404)


@login_required
def student_progress_view(request):
    user = request.user
    if user.role != 'student':
        return HttpResponseForbidden("Доступ только для студентов.")

    student_group = user.group
    if not student_group:
        return render(request, 'main/student_progress.html', {'error': "Вы не прикреплены к группе."})

    assignments = GroupDisciplineAssignment.objects.filter(group=student_group).select_related('curriculum_item__subject')
    student_works = StudentWork.objects.filter(student=user)

    works_by_assignment = defaultdict(list)
    for work in student_works:
        works_by_assignment[work.assignment_id].append(work)

    progress_data = defaultdict(lambda: defaultdict(list))

    for assignment in assignments:
        curriculum = assignment.curriculum_item
        subject = curriculum.subject.name
        course = curriculum.course
        semester = curriculum.semester
        works = works_by_assignment.get(assignment.id, [])

        # Лабы
        labs = []
        for i in range(assignment.lab_count):
            lab = next((w for w in works if w.work_type == 'lab' and w.lab_index == i), None)
            labs.append({'label': f'Лаб. {i+1}', 'status': '✅' if lab and lab.is_completed else '❌'})

        # Практики
        practices = []
        for i in range(assignment.practice_count):
            pr = next((w for w in works if w.work_type == 'practice' and w.lab_index == i), None)
            label = f'Пр. {i+1}'
            status = str(pr.grade) if pr and pr.grade else '❌'
            practices.append({'label': label, 'status': status})

        # Прочие работы (курсовые, ргр и т.д.)
        work_types = curriculum.work_types.split(', ') if curriculum.work_types else []
        existing_other = [w for w in works if w.work_type in work_types]

        other_works = []
        for wt in work_types:
            matching = [w for w in existing_other if w.work_type == wt]
            if matching:
                w = matching[0]
                status = str(w.grade) if w.grade else ('✅' if w.is_passed else '❌')
            else:
                status = '❌'
            other_works.append({'label': wt, 'status': status})

        # Контроль (зачет, экзамен и т.д.)
        control_types = curriculum.control_types.split(', ') if curriculum.control_types else []
        existing_control = [w for w in works if w.work_type in control_types]

        control_works = []
        for ct in control_types:
            matching = [w for w in existing_control if w.work_type == ct]
            if matching:
                w = matching[0]
                status = str(w.grade) if w.grade else ('✅' if w.is_passed else '❌')
            else:
                status = '❌'
            control_works.append({'label': ct, 'status': status})
        total = len(labs) + len(practices) + len(other_works) + len(control_works)
        done = sum(1 for lab in labs if lab['status'] == '✅') + \
               sum(1 for p in practices if p['status'] != '❌') + \
               sum(1 for w in other_works if w['status'] != '❌') + \
               sum(1 for w in control_works if w['status'] != '❌')

        progress = int((done / total) * 100) if total else 0

        progress_data[course][semester].append({
            'subject': subject,
            'labs': labs,
            'practices': practices,
            'other_works': other_works,
            'control_works': control_works,
            'progress': progress,
        })

    # Сортируем и рассчитываем средний прогресс
    structured_data = []
    for course in sorted(progress_data):
        semesters = []
        course_total_progress = []
        for semester in sorted(progress_data[course]):
            disciplines = progress_data[course][semester]
            semester_progress = [d['progress'] for d in disciplines]
            avg_semester_progress = int(sum(semester_progress) / len(semester_progress)) if semester_progress else 0
            course_total_progress.extend(semester_progress)

            semesters.append({
                'semester': semester,
                'disciplines': disciplines,
                'avg_progress': avg_semester_progress,  # ← добавлено
            })

        avg_course_progress = int(
            sum(course_total_progress) / len(course_total_progress)) if course_total_progress else 0

        structured_data.append({
            'course': course,
            'semesters': semesters,
            'avg_progress': avg_course_progress,  # ← добавлено
        })

    context = {
        'progress_data': structured_data
    }

    return render(request, 'main/student_progress.html', context)